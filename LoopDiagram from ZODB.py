# -*- coding: utf-8 -*-
"""
Created on Wed May 17 14:16:46 2023

@author: MMAR
"""

from Classes.DS_BES_Cable import DS_BES_Cable
from Classes.DS_BES_Slipring import DS_BES_Slipring
from ZEO import ClientStorage
from ZODB import FileStorage, DB

from collections import defaultdict
import os
import openpyxl
from openpyxl.styles import Alignment, PatternFill

addr = '10.175.13.199', 8091 # This is the address of the ZEO 


def generate_excel(cables_sorted_by_signal):
    # Name of the file
    excel_file = "cable_connections.xlsx"

    # Check if the file exists.
    if os.path.exists(excel_file):
        # Delete if exists the excel file
        os.remove(excel_file)
    
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    row = 2  # Start from the second row
    for signal, cables_ordenados in cables_sorted_by_signal.items():
        # Write the signal name in the current row
        sheet.cell(row=row, column=1, value=signal)
            
        # Write the cable connections and cores
        connection_column = 2  # Start from the second column
        
        for i, key in enumerate(cables_ordenados):
            if i == 0:
                # Write the connection information in the row above
                sheet.cell(row=row-1, column=connection_column, value=CableDict[key][2])    #connection 1
                sheet.cell(row=row-1, column=connection_column+1, value=key).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")   #cable Tag in yellow
                sheet.cell(row=row-1, column=connection_column+2, value=CableDict[key][4])  #connection 2
                
                r = row -1
                
                for j, current_signal in enumerate(CableDict[key][1]): #signals
                    if signal == current_signal: 
                        r += 1
                        # Write the position information in the current row
                        sheet.cell(row=r, column=connection_column, value=CableDict[key][3][j])   #position connection 1
                        sheet.cell(row=r, column=connection_column+2, value=CableDict[key][5][j]) #position connection 2
                
                # Write the conductor in the row below
                        sheet.cell(row=r, column=connection_column+1, value= str(CableDict[key][0][j])+ ' (' + str(CableDict[key][6][j]) + ')')      #core and color
                
                
                 # Write the signal name in the current row
                        sheet.cell(row=r, column=1, value=signal)

                # Increment the connection column for the next cable
                connection_column += 3
                
                # If the fist cable is longer than the rest, the last row should be this one
                last_row = r
                
            else:
                if CableDict[key][2] in SlipRingTags:  #if there is a Slip Ring in the connection 1, should be added. 
                    r = row -1
                    sheet.cell(row=row-1, column=connection_column, value=CableDict[key][2])    #connection 1
                    for j, current_signal in enumerate(CableDict[key][1]): #signals
                        if signal == current_signal: 
                            r += 1
                            # Write the position information in the current row
                            sheet.cell(row=r, column=connection_column, value=CableDict[key][3][j])   #position connection 1
                    connection_column += 1
                    
                sheet.cell(row=row-1, column=connection_column, value=key).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")   #cable Tag in yellow 
                sheet.cell(row=row-1, column=connection_column+1, value=CableDict[key][4]) #connection 2
                r = row -1
                
                for j, current_signal in enumerate(CableDict[key][1]):
                    if signal == current_signal:
                        r += 1
                # Write the position information in the current row
                        sheet.cell(row=r, column=connection_column+1, value=CableDict[key][5][j])   #position connection 2
                
                # Write the conductor in the row below
                        sheet.cell(row=r, column=connection_column, value= str(CableDict[key][0][j])+ ' (' + str(CableDict[key][6][j]) + ')')       #core and color
                
                # Increment the connection column for the next cable
                connection_column += 2
            
        row= last_row

        # Increment the row for the next signal
        row += 4  # Increase by 4 to leave two blank rows between signals
    
    # Set the column width to 15 and align the content to the center
    for column in sheet.columns:
        column_width = 25
        column = tuple(column)
        for cell in column:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions[column[0].column_letter].width = column_width
    
    
    # Save the workbook as an Excel file
    workbook.save(excel_file)
    print("Excel file generated successfully.")


CableDict = {}
SlipRingTags = []
 
try:
        storage = ClientStorage.ClientStorage(addr)             
        db = DB(storage)
        connection = db.open()
        root = connection.root()
        for key in root: 
            obj = root[key]                
                
            if isinstance(obj,DS_BES_Slipring):
                SlipRingTags.append(key)

            if isinstance(obj,DS_BES_Cable):
                # if obj.TagNumber in ["IC-SPM-IJB-7507-02","IC-SPM-IJB-6650-01"]:   #To show all the attributes of a selected cables
                #     # Get all attributes and values of the object
                #     attributes = vars(obj)

                #     # Print the attributes and values
                #     for attribute, value in attributes.items():
                #         print(f"{attribute}: {value}")
                
                CableDict[key] = []
                cores  = []
                signals = []
                connection1_positions = []
                connection2_positions = []
                colors = []
                
                for i in range(1,41): # maximum number of cores is 41
                    if getattr(obj,"CoreNumber"+str(i)) != "":
                        cores.append(getattr(obj,"CoreNumber"+str(i)))
                        
                    if getattr(obj,"LSignal"+str(i)) != "" and getattr(obj,"LSignal"+str(i)) != None:
                        signals.append(getattr(obj,"LSignal"+str(i)))
                        
                    if getattr(obj,"LTermination"+str(i)) != "":
                        connection1_positions.append(getattr(obj,"LTermination"+str(i)))
                                    
                    if getattr(obj,"RTermination"+str(i)) != "":
                        connection2_positions.append(getattr(obj,"RTermination"+str(i)))
                
                    if getattr(obj,"Color"+str(i)) != "":
                        colors.append(getattr(obj,"Color"+str(i)))

                CableDict[key].append(cores) 
                CableDict[key].append(signals) 
                CableDict[key].append(obj.Connection1)
                CableDict[key].append(connection1_positions) 
                CableDict[key].append(obj.Connection2)
                CableDict[key].append(connection2_positions) 
                CableDict[key].append(colors) 

            # if key in ["IC-SPM-IJB-7507-02","IC-SPM-IJB-6650-01"]:      #To see the cable dict details
            #     print(key)
            #     print(CableDict[key])

        connection.close()
        storage.close()
                    
#                    -----------------------------------------------------
        # Separate the cables by signal 
        allSignals = []
        cables_by_signal = defaultdict(list)
        for key in CableDict:
#            print(CableDict[key][1])
            for signal in CableDict[key][1]:
                allSignals.append(signal)
                if key not in cables_by_signal[signal]:
                    cables_by_signal[signal].append(key)
        allSignals = list(dict.fromkeys(allSignals))

        #                    -----------------------------------------------------               
        excluded_signals = ['SPM-TIT-7101','SPM-TIT-7115','SPM-PSLL-7110B','SPM-XL-7110A','SPM-PSLL-7110A','SPM-HPU-9401-SW10','SPM-HPU-9401-SW11','SPARE','F.O. PCS','F.O. PCS (SPARE)','F.O. SIS','F.O. SIS (SPARE)','POWER +','POWER -','RS-485 DATA +','RS-485 DATA -','4-20 mA +','4-20 mA -','RELAY Vo +','RELAY Vo -','RESET','-','ETHERNET','REDUNDANT WIRE','INDICATOR','INSTR. EARTH']
        cables_sorted_by_signal = {}
        for signal, keys in cables_by_signal.items():
            # if signal in ["SPM-LIT-6651","SPM-PIT-6645","SPM-HS-6602A","SPM-HS-3501A"]:
            if signal in list(dict.fromkeys(allSignals)) and signal not in excluded_signals:
                # print (signal)
                used_connections = set()
                initial_cable = None
                for key in keys:
                    used_connections.add(CableDict[key][4])
                    
                for key in keys:
                    if CableDict[key][2] not in used_connections and CableDict[key][2] not in SlipRingTags:
                        initial_cable = key

                sorted_cables = [initial_cable]

                counter = 0
                while len(sorted_cables) < len(keys):
                    if counter > 50: 
                        print(f"There is a problem with the follwing signal {signal}")
                        break
                    for key in keys:
                        if key not in sorted_cables and CableDict[key][2] == CableDict[sorted_cables[-1]][4]:
                            sorted_cables.append(key)
                    for key in keys:
                        if key not in sorted_cables and CableDict[key][2] in SlipRingTags:
                            sorted_cables.append(key)
                    counter += 1
                        
                cables_sorted_by_signal[signal] = sorted_cables
#                    -----------------------------------------------------
        
        # Call the function with the cables_sorted_by_signal dictionary as input
        #print(cables_sorted_by_signal)
        generate_excel(cables_sorted_by_signal)
        
except BaseException as e:
    # Code to handle the exception
        print("An exception occurred:", e)
        connection.close()
        storage.close()  
